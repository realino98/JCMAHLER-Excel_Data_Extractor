import pandas as pd
import xlwt
import os
from openpyxl import load_workbook

EXCEL_FILE_NAME = "Referrals.xlsx"
FIRST_SHEET_NAME = "Work Received inc amavat 2021"
SECOND_SHEET_NAME = "Work Referred inc amavat 2021"
OUTPUT = "Output/"

df = pd.read_excel('Files/Referrals.xlsx', sheet_name=[FIRST_SHEET_NAME, SECOND_SHEET_NAME])
df1 = df[FIRST_SHEET_NAME]
df2 = df[SECOND_SHEET_NAME]

print("Getting all the company Names...")
companies_1 = []
# print(len(df1["To Company"]))
for i in range(len(df1["To Company"])):
    # print(i)
    try:
        if df1["To Company"][i] not in companies_1 and "Total" not in df1["To Company"][i]:
            companies_1.append(df1["To Company"][i])
    except:
        pass

# print(companies_1, len(companies_1))

companies_2 = []
# print(len(df2["From Company"]))
for i in range(len(df2["From Company"])):
    # print(i)
    try:
        if df2["From Company"][i] not in companies_2 and "Total" not in df2["From Company"][i]:
            companies_2.append(df2["From Company"][i])
    except:
        pass

# print(companies_2, len(companies_2))
print("Generating Files...")
for i in range(len(companies_1)):
# for i in range(5):
    # print(i)
    path = OUTPUT+companies_1[i].replace("/","_")+".xlsx"
    new_df_1 = df1.loc[df1["To Company"].isin([companies_1[i], companies_1[i]+" Total"])]
    print("Creating Sheet 1", companies_1[i]+".xlsx")
    writer = pd.ExcelWriter(path, engine = 'xlsxwriter')
    new_df_1.to_excel(writer, index=False, sheet_name = FIRST_SHEET_NAME)
    writer.save()
    writer.close()
    # print(new_df_1)

for i in range(len(companies_2)):
# for i in range(5):
    # print(i)
    path = OUTPUT+companies_2[i].replace("/","_")+".xlsx"
    new_df_2 = df2.loc[df2["From Company"].isin([companies_2[i], companies_2[i]+" Total"])]
    print("Creating Sheet 2", companies_2[i]+".xlsx")
    try: 
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
    except:
        writer = pd.ExcelWriter(path, engine = 'xlsxwriter')
    new_df_2.to_excel(writer, index=False, sheet_name = SECOND_SHEET_NAME)
    writer.save()
    writer.close()
    # print(new_df_2)

print("Excel File Generated to Output/..")
from openpyxl import load_workbook

outputs = os.listdir("Output/")
for output in outputs:
    try:
        wb = load_workbook(OUTPUT+output)
        ws = wb[FIRST_SHEET_NAME]
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 30
        ws.column_dimensions["I"].width = 30
        ws.column_dimensions["J"].width = 30
        ws.column_dimensions["K"].width = 30
        ws.column_dimensions["L"].width = 30
        ws.column_dimensions["M"].width = 30
        wb.save(OUTPUT+output)
        wb.close()
    except:
        pass

for output in outputs:
    try:
        wb = load_workbook(OUTPUT+output)
        ws = wb[SECOND_SHEET_NAME]
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 30
        ws.column_dimensions["I"].width = 30
        ws.column_dimensions["J"].width = 30
        ws.column_dimensions["K"].width = 30
        ws.column_dimensions["L"].width = 30
        ws.column_dimensions["M"].width = 30
        wb.save(OUTPUT+output)
        wb.close()
    except:
        pass